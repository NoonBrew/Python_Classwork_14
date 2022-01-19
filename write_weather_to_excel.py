from openpyxl import Workbook

week_temps = {
    'Monday': 54,
    'Tuesday': 60,
    'Wednesday': 62,
    'Thursday': 57,
    'Friday': 71,
}

workbook = Workbook()

worksheet = workbook.active

row_index = 2
worksheet.cell(1,1, 'Day')
worksheet.cell(1,2, 'Temperature (F)')
worksheet.title = 'Daily Temperatures'


for day, temp in week_temps.items():
    worksheet.cell(row_index, 1, day)
    worksheet.cell(row_index, 2, temp)
    row_index += 1

workbook.save('Temperatures.xlsx')
