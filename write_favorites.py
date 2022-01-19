from openpyxl import Workbook

favorite_foods = ['Pizza', 'Cheese Burgers', 'Ice Cream']
favorite_colors = ['Green', 'Light Blue', 'Dark Red', 'Cyan']

workbook = Workbook()

worksheet = workbook.active

worksheet.title = 'Favorite Things'

worksheet.cell(1,1, 'Favorite Foods')
# index + 1 writes to row, 1 writes to column, food is entry

for index, food in enumerate(favorite_foods):
    worksheet.cell(index+2, 1, food)

worksheet.cell(1,2, 'Favorite Colors')
for index, color in enumerate(favorite_colors):
    worksheet.cell(index+2, 2, color)

workbook.save('favorites.xlsx')